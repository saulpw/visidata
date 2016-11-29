#!/usr/bin/env python3
#
# VisiData: a curses interface for exploring and arranging tabular data
#
# Copyright (C) 2016 Saul Pwanson
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this program.  If not, see <http://www.gnu.org/licenses/>.

__version__ = '0.31'
__author__ = 'Saul Pwanson <vd@saul.pw>'
__license__ = 'GPLv3'
__status__ = 'Development'

import itertools
import os.path
import io
import datetime
import string
import collections
import functools
import codecs
import statistics
import curses
import curses.textpad
import re
import html.parser
import urllib.parse
import csv

initialStatus = 'saul.pw/VisiData v' + __version__
defaultColNames = list(itertools.chain(string.ascii_uppercase, [''.join(i) for i in itertools.product(string.ascii_uppercase, repeat=2)]))

default_options = {
    'csv_dialect': 'excel',
    'csv_delimiter': ',',
    'csv_quotechar': '"',
    'csv_header': False,

    'debug': False,
    'readonly': False,

    'encoding': 'utf-8',
    'encoding_errors': 'surrogateescape',

    'VisibleNone': '',   # visible contents of a cell whose value was None
    'ColumnFiller': ' ',
    'ColumnSep': ' | ',  # chars between columns
    'Ellipsis': '…',
    'SubsheetSep': '~',
    'StatusSep': ' | ',
    'KeySep': '/',
    'SheetNameFmt': '%s| ',  # before status line
    'FunctionError': '¿',    # when computation fails due to exception
    'HistogramChar': '*',
    'ColumnStats': False,  # whether to include mean/median/etc on 'C'olumn sheet
    'EditPadChar': '_',
    'Unprintable': '.',

    # color scheme
    'c_default': 'normal',
    'c_Header': 'bold',
    'c_CurHdr': 'reverse',
    'c_CurRow': 'reverse',
    'c_CurCol': 'bold',
    'c_KeyCols': 'brown',
    'c_StatusLine': 'bold',
    'c_SelectedRow': 'green',
    'c_ColumnSep': 'blue',
    'c_EditCell': 'normal',
}

vd = None    # toplevel VisiData, contains all sheets
scr = None   # toplevel curses screen
sheet = None # current sheet

windowWidth = None
windowHeight = None

colors = {
    'bold': curses.A_BOLD,
    'reverse': curses.A_REVERSE,
    'normal': curses.A_NORMAL,
}


def keyname(ch):
    return curses.keyname(ch).decode('utf-8')

class attrdict(object):
    def __init__(self, d):
        self.__dict__ = d

options = attrdict(default_options)

class VException(Exception):
    pass

class VEscape(VException):
    pass

class ChangeCommandSet(VException):
    def __init__(self, commands, mode):
        self.commands = commands
        self.mode = mode

def ctrl(ch):
    return ord(ch) & 31  # convert from 'a' to ^A keycode

ENTER = ctrl('j')
ESC = 27

base_commands = {}
global_commands = {}
def command(ch, cmdstr, helpstr=''):
    base_commands[ch] = (cmdstr, helpstr)

def global_command(ch, cmdstr, helpstr):
    global_commands[ch] = (cmdstr, helpstr)

command(curses.KEY_F1,    'vd.push(sheet.CommandHelp())', 'command help')
command(ord('q'),         'vd.sheets.pop(0)', 'quit current sheet')

command(curses.KEY_LEFT,  'sheet.moveCursorRight(-1)', 'arrow keys move cursor')
command(curses.KEY_DOWN,  'sheet.moveCursorDown(+1)')
command(curses.KEY_UP,    'sheet.moveCursorDown(-1)')
command(curses.KEY_RIGHT, 'sheet.moveCursorRight(+1)')
command(curses.KEY_NPAGE, 'sheet.moveCursorDown(sheet.nVisibleRows); sheet.topRowIndex += sheet.nVisibleRows', 'scroll down one page')
command(curses.KEY_PPAGE, 'sheet.moveCursorDown(-sheet.nVisibleRows); sheet.topRowIndex -= sheet.nVisibleRows', 'scroll up one page')
command(curses.KEY_HOME,  'sheet.topRowIndex = sheet.cursorRowIndex = 0', 'go to top row')
command(curses.KEY_END,   'sheet.cursorRowIndex = len(sheet.rows)-1', 'go to last row')

command(ord('h'), 'sheet.moveCursorRight(-1)', 'hjkl keys move cursor')
command(ord('j'), 'sheet.moveCursorDown(+1)')
command(ord('k'), 'sheet.moveCursorDown(-1)')
command(ord('l'), 'sheet.moveCursorRight(+1)')

command(ord('H'), 'sheet.cursorColIndex = moveListItem(sheet.columns, sheet.cursorColIndex, max(sheet.cursorColIndex-1, 0))', 'shift-movement reorders current row/column')
command(ord('J'), 'sheet.cursorRowIndex = moveListItem(sheet.rows, sheet.cursorRowIndex, min(sheet.cursorRowIndex+1, sheet.nRows-1))')
command(ord('K'), 'sheet.cursorRowIndex = moveListItem(sheet.rows, sheet.cursorRowIndex, max(sheet.cursorRowIndex-1, 0))')
command(ord('L'), 'sheet.cursorColIndex = moveListItem(sheet.columns, sheet.cursorColIndex, min(sheet.cursorColIndex+1, sheet.nCols))')

command(ctrl('g'), 'vd.status(sheet.statusLine)', 'current sheet info')
command(ctrl('p'), 'vd.status(vd.statusHistory[0])', 'show previous status line again')
command(ctrl('v'), 'vd.status(initialStatus)', 'get version information')

command(ord('t'), 'sheet.topRowIndex = sheet.cursorRowIndex', 'scroll cursor row to top of screen')
command(ord('m'), 'sheet.topRowIndex = sheet.cursorRowIndex-int(sheet.nVisibleRows/2)', 'scroll cursor row to middle of screen')
command(ord('b'), 'sheet.topRowIndex = sheet.cursorRowIndex-sheet.nVisibleRows+1', 'scroll cursor row to bottom of screen')

command(ord('<'), 'sheet.skipUp()', 'skip up current column to previous value')
command(ord('>'), 'sheet.skipDown()', 'skip down current column to next value')

command(ord('_'), 'sheet.cursorCol.width = getMaxWidth(sheet.cursorCol, sheet.visibleRows)', 'set column width to max visible cell on screen')
command(ord('-'), 'sheet.columns.pop(sheet.cursorColIndex)', 'delete current column')
command(ord('^'), 'sheet.cursorCol.name = sheet.cursorCol.getDisplayValue(sheet.cursorRow)', 'set current column header to current cell value')
command(ord('!'), 'sheet.toggleKeyColumn(sheet.cursorColIndex)', 'toggle this column as key')

# command(ord('@'), 'sheet.convertType(sheet.cursorCol, datetime)', 'convert column to ISO8601 date')
command(ord('#'), 'sheet.convertType(sheet.cursorCol, int)', 'convert column to integer')
command(ord('$'), 'sheet.convertType(sheet.cursorCol, str)', 'convert column to string')
command(ord('%'), 'sheet.convertType(sheet.cursorCol, float)', 'convert column to decimal numeric type')
command(ord('['), 'sheet.rows = sorted(sheet.rows, key=lambda r, sheet.cursorCol.getValue(r) or sheet.cursorCol.type())', 'sort by current column ascending')
command(ord(']'), 'sheet.rows = sorted(sheet.rows, key=lambda r, sheet.cursorCol.getValue(r) or sheet.cursorCol.type(), reverse=True)', 'sort by current column descending')
command(ctrl('e'), 'options.debug = True; raise VException(vd.lastErrors[-1])', 'quit and print error to terminal')
command(ctrl('d'), 'options.debug = not options.debug; vd.status("debug " + ("ON" if options.debug else "OFF"))', 'toggle debug option')

command(ord('E'), 'if vd.lastErrors: vd.push(VSheetText("last_error", vd.lastErrors[-1]))', 'display stack trace for most recent exception')
command(ord('F'), 'vd.push(VSheetFreqTable(sheet, sheet.cursorCol))', 'create frequency table from values in current column')

command(ord('d'), 'sheet.rows.pop(sheet.cursorRowIndex)', 'delete current row')

command(ord('g'), 'raise ChangeCommandSet(global_commands, "g")', 'global prefix')

command(ord('S'), 'vd.push(vd.sheets)', 'push sheets sheet')
command(ord('C'), 'vd.push(VSheetColumns(sheet))', 'push columns sheet')
command(ord('O'), 'vd.push(VSheetObject("options", options))', 'push options sheet')

command(ord('/'), 'sheet.searchRegex(inputLine(prompt="/"), columns=[sheet.cursorCol], moveCursor=True)', 'search this column forward for regex')
command(ord('?'), 'sheet.searchRegex(inputLine(prompt="?"), columns=[sheet.cursorCol], backward=True, moveCursor=True)', 'search this column backward for regex')
command(ord('n'), 'sheet.searchRegex(columns=[sheet.cursorCol], moveCursor=True)', 'repeat last search forward')
command(ord('p'), 'sheet.searchRegex(columns=[sheet.cursorCol], backward=True, moveCursor=True)', 'repeat last search backward')

command(ord(' '), 'sheet.toggle([sheet.cursorRow]); sheet.moveCursorDown(1)', 'toggle select of current row')
command(ord('s'), 'sheet.select([sheet.cursorRow]); sheet.moveCursorDown(1)', 'select current row')
command(ord('u'), 'sheet.unselect([sheet.cursorRow]); sheet.moveCursorDown(1)', 'unselect current row')
command(ord('|'), 'sheet.select(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="|"), columns=[sheet.cursorCol]))', 'select rows by regex in this column')
command(ord('\\'), 'sheet.unselect(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="\\\\"), columns=[sheet.cursorCol]))', 'unselect rows by regex in this column')

command(ord('R'), 'sheet.source.type = inputLine("change type to, ") or sheet.source.type', 'set parsing type of current sheet')
command(ctrl('r'), 'openFileOrUrl(vd.sheets.pop(0).source); vd.status("reloaded")', 'reload current sheet')
command(ctrl('s'), 'saveSheet(sheet, inputLine("save to, "))', 'save sheet (type determined by ext)')
command(ord('o'), 'openFileOrUrl(inputLine("open: "))', 'open a file or url')
command(ctrl('o'), 'expr = inputLine("eval: "); pushPyObjSheet(expr, eval(expr))', 'eval an expression and show the result')

command(ord('e'), 'sheet.cursorCol.setValue(sheet.cursorRow, sheet.editCell(sheet.cursorColIndex))', 'edit current cell')

command(ord('='), 'sheet.addColumn(ColumnExpr(sheet, inputLine("=")))', 'add column by expr')
command(ctrl('^'), 'vd.sheets[0], vd.sheets[1] = vd.sheets[1], vd.sheets[0]', 'swap top two sheets (like vi)')

# when used with 'g' prefix
global_command(ord('q'), 'vd.sheets.clear()', 'quit all sheets (and therefore exit)')

global_command(ord('h'), 'sheet.cursorColIndex = sheet.leftColIndex = 0', 'move cursor to leftmost column')
global_command(ord('k'), 'sheet.cursorRowIndex = sheet.topRowIndex = 0', 'move cursor to top row')
global_command(ord('j'), 'sheet.cursorRowIndex = len(sheet.rows); sheet.topRowIndex = sheet.cursorRowIndex-sheet.nVisibleRows', 'move cursor to bottom row')
global_command(ord('l'), 'sheet.cursorColIndex = len(sheet.columns)-1', 'move cursor to rightmost column')

global_command(ord('H'), 'moveListItem(sheet.columns, sheet.cursorColIndex, 0)', 'throw current column all the way to the left')
global_command(ord('J'), 'moveListItem(sheet.rows, sheet.cursorRowIndex, sheet.nRows)', 'throw current row all the way to the bottom')
global_command(ord('K'), 'moveListItem(sheet.rows, sheet.cursorRowIndex, 0)', 'throw current row all the way to the top')
global_command(ord('L'), 'moveListItem(sheet.columns, sheet.cursorColIndex, sheet.nCols)', 'throw current column all the way to the right')

global_command(ord('_'), 'for c in sheet.columns: c.width = getMaxWidth(c, sheet.visibleRows)', 'resize all columns to max width of visible cells on screen')
global_command(ord('^'), 'for c in sheet.columns: c.name = c.getDisplayValue(sheet.cursorRow)', 'set all column names to values in the current row')

global_command(ord('E'), 'vd.push(VSheetText("last_error", "\\n\\n".join(vd.lastErrors)))', 'push sheet of all previous errors')

global_command(ord('/'), 'sheet.searchRegex(inputLine(prompt="/"), moveCursor=True, columns=sheet.columns)', 'search regex forward in all columns')
global_command(ord('?'), 'sheet.searchRegex(inputLine(prompt="?"), backward=True, moveCursor=True, columns=sheet.columns)', 'search regex backward in all columns')
global_command(ord('n'), 'sheet.cursorRowIndex = max(sheet.searchRegex())', 'move cursor to first match')
global_command(ord('p'), 'sheet.cursorRowIndex = min(sheet.searchRegex())', 'move cursor to last match')

global_command(ord(' '), 'sheet.toggle(sheet.rows)', 'toggle all rows')
global_command(ord('s'), 'sheet.select(sheet.rows)', 'select all rows')
global_command(ord('u'), 'sheet._selectedRows = {}', 'unselect all rows')

global_command(ord('|'), 'sheet.select(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="|"), columns=sheet.columns))', 'select rows by regex in all columns')
global_command(ord('\\'), 'sheet.unselect(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="\\\\"), columns=sheet.columns))', 'unselect rows by regex in all columns')

global_command(ord('d'), 'sheet.rows = [r for r in sheet.rows if not sheet.isSelected(r)]; sheet._selectedRows = {}', 'delete all selected rows')

global_command(ctrl('p'), 'vd.push(VSheetText("statuses", vd.statusHistory))', 'open sheet with all previous messages')

### VisiData core

def moveListItem(L, fromidx, toidx):
    r = L.pop(fromidx)
    L.insert(toidx, r)
    return toidx

def iso8601(dt):
    return dt.strftime('%Y-%m-%d %H:%M:%S')

def datestr(t):
    return iso8601(datetime.datetime.fromtimestamp(t))

def getMaxWidth(col, rows):
    if len(rows) == 0:
        return 0

    return min(max(max(len(col.getDisplayValue(r)) for r in rows), len(col.name)+2), int(windowWidth-1))


class VSheet:
    def __init__(self, name, src=None):
        self.name = name
        self.source = src
        self.rows = []
        self.cursorRowIndex = 0  # absolute index of cursor into self.rows
        self.cursorColIndex = 0  # absolute index of cursor into self.columns

        self.topRowIndex = 0     # cursorRowIndex of topmost row
        self.leftColIndex = 0    # cursorColIndex of leftmost column

        # as computed during draw()
        self.rowLayout = {}      # [rowidx] -> y
        self.colLayout = {}      # [colidx] -> (x, w)

        # all columns in display order
        self.columns = [ ]
        self.nKeys = 0           # self.columns[:nKeys] are all pinned to the left and matched on join

        # current search term
        self.currentRegex = None
        self.currentRegexColumns = None

        self._selectedRows = {}   # id(row) -> row

        # specialized sheet keys
        self.commands = base_commands.copy()

    def __repr__(self):
        return self.name

    def isSelected(self, r):
        return id(r) in self._selectedRows

    def command(self, key, cmdstr, helpstr=''):
#        if key in self.commands:
#            vd.status('overriding key %s' % key)
        self.commands[key] = (cmdstr, helpstr)

    @property
    def nVisibleRows(self):
        return windowHeight-2

    @property
    def cursorCol(self):
        return self.columns[self.cursorColIndex]

    @property
    def cursorRow(self):
        return self.rows[self.cursorRowIndex]

    @property
    def visibleRows(self):
        return self.rows[self.topRowIndex:self.topRowIndex+windowHeight-2]

    @property
    def selectedRows(self):
        return [r for r in self.rows if id(r) in self._selectedRows]

    @property
    def keyCols(self):
        return self.columns[:self.nKeys]

    @property
    def keyColNames(self):
        return options.KeySep.join(c.name for c in self.keyCols)

    @property
    def cursorValue(self):
        return self.cellValue(self.cursorRowIndex, self.cursorColIndex)

    @property
    def statusLine(self):
        return 'row %s/%s' % (self.cursorRowIndex, len(self.rows))

    @property
    def nRows(self):
        return len(self.rows)

    @property
    def nCols(self):
        return len(self.columns)

    def moveCursorDown(self, n):
        self.cursorRowIndex += n

    def moveCursorRight(self, n):
        self.cursorColIndex += n

    def cellValue(self, rownum, col):
        if not isinstance(col, VColumn):
            # assume it's the column number
            col = self.columns[col]
        return col.getValue(self.rows[rownum])

    @functools.lru_cache()
    def columnValues(self, col):
        if not isinstance(col, VColumn):
            # assume it's the column number
            col = self.columns[col]
        return [col.getValue(r) for r in self.rows]

    def addColumn(self, col):
        if col:
            self.columns.append(col)

    def toggleKeyColumn(self, colidx):
        if self.cursorColIndex >= self.nKeys: # if not a key, add it
            moveListItem(self.columns, self.cursorColIndex, self.nKeys)
            self.nKeys += 1
        else:  # otherwise move it after the last key
            self.nKeys -= 1
            moveListItem(self.columns, self.cursorColIndex, self.nKeys)

    def skipDown(self):
        pv = self.cursorValue
        for i in range(self.cursorRowIndex+1, len(self.rows)):
            if self.cellValue(i, self.cursorColIndex) != pv:
                self.cursorRowIndex = i
                return

        vd.status('no different value down this column')

    def skipUp(self):
        pv = self.cursorValue
        for i in range(self.cursorRowIndex, -1, -1):
            if self.cellValue(i, self.cursorColIndex) != pv:
                self.cursorRowIndex = i
                return

        vd.status('no different value up this column')

    def convertType(self, col, newType):
        nErrors = 0
        for r in self.rows:
            try:
                col.setValue(r, newType(col.getValue(r)))
            except Exception as e:
                col.setValue(r, None)
                nErrors += 1
        col.type = newType
        vd.status('converted %s to %s with %s exceptions' % (col.name, newType.__name__, nErrors))

    def toggle(self, rows):
        for r in rows:
            if id(r) in self._selectedRows:
                del self._selectedRows[id(r)]
            else:
                self._selectedRows[id(r)] = r

    def select(self, rows):
        rows = list(rows)
        before = len(self._selectedRows)
        self._selectedRows.update(dict((id(r), r) for r in rows))
        vd.status('selected %s/%s rows' % (len(self._selectedRows)-before, len(rows)))

    def unselect(self, rows):
        rows = list(rows)
        before = len(self._selectedRows)
        for r in rows:
            del self._selectedRows[id(r)]
        vd.status('unselected %s/%s rows' % (before-len(self._selectedRows), len(rows)))

    def columnsMatch(self, row, columns, func):
        for c in columns:
            m = func(c.getDisplayValue(row))
            if m:
                return True
        return False

    def checkCursor(self):
        # keep cursor within actual available rowset
        if self.cursorRowIndex <= 0:
            self.cursorRowIndex = 0
        elif self.cursorRowIndex >= len(self.rows):
            self.cursorRowIndex = len(self.rows)-1

        if self.cursorColIndex <= 0:
            self.cursorColIndex = 0
        elif self.cursorColIndex >= len(self.columns):
            self.cursorColIndex = len(self.columns)-1

        if self.topRowIndex <= 0:
            self.topRowIndex = 0
        elif self.topRowIndex > len(self.rows):
            self.topRowIndex = len(self.rows)-1

        # (x,y) is relative cell within screen viewport
        x = self.cursorColIndex - self.leftColIndex
        y = self.cursorRowIndex - self.topRowIndex + 1  # header

        # check bounds, scroll if necessary
        if y < 1:
            self.topRowIndex = self.cursorRowIndex
        elif y > windowHeight-2:
            self.topRowIndex = self.cursorRowIndex-windowHeight+3

        if x <= 0:
            self.leftColIndex = self.cursorColIndex
        else:
            # keycolwidth = sum(self.columns[i].width for i in range(0, self.nKeys))
            while True:
                self.calcColLayout()
                if self.cursorColIndex not in self.colLayout:  # should always be to the right
                    assert self.cursorColIndex > self.leftColIndex
                    self.leftColIndex += 1
                    continue
                cur_x, cur_w = self.colLayout[self.cursorColIndex]
                left_x, left_w = self.colLayout[self.leftColIndex]
                if cur_x+cur_w < windowWidth:
                    # current columns fit entirely on screen
                    break
                self.leftColIndex += 1

    def searchRegex(self, regex=None, columns=None, backward=False, moveCursor=False):
        'sets row index if moveCursor; otherwise returns list of row indexes'
        if regex:
            self.currentRegex = re.compile(regex, re.IGNORECASE)

        if not self.currentRegex:
            vd.status('no regex')
            return []

        if columns:
            self.currentRegexColumns = columns

        if not self.currentRegexColumns:
            vd.status('no columns given')
            return []

        if backward:
            rng = range(self.cursorRowIndex-1, -1, -1)
            rng2 = range(self.nRows-1, self.cursorRowIndex-1, -1)
        else:
            rng = range(self.cursorRowIndex+1, self.nRows)
            rng2 = range(0, self.cursorRowIndex+1)

        matchingRowIndexes = []

        for r in rng:
            if self.columnsMatch(self.rows[r], self.currentRegexColumns, self.currentRegex.search):
                if moveCursor:
                    self.cursorRowIndex = r
                    return r
                matchingRowIndexes.append(r)

        for r in rng2:
            if self.columnsMatch(self.rows[r], self.currentRegexColumns, self.currentRegex.search):
                if moveCursor:
                    self.cursorRowIndex = r
                    vd.status('search wrapped')
                    return r
                matchingRowIndexes.append(r)

        vd.status('%s matches for /%s/' % (len(matchingRowIndexes), self.currentRegex.pattern))

        return matchingRowIndexes

    def calcColLayout(self):
        self.colLayout = {}
        x = 0
        for colidx in range(0, len(self.columns)):
            col = self.columns[colidx]
            col.width = col.width or getMaxWidth(col, self.visibleRows)
            if colidx < self.nKeys or colidx >= self.leftColIndex:  # visible columns
                self.colLayout[colidx] = (x, min(col.width, windowWidth-x))
                x += col.width+len(options.ColumnSep)
            if x > windowWidth:
                break

    def draw(self):
        scr.erase()  # clear screen before every re-draw
        sepchars = options.ColumnSep

        self.rowLayout = {}
        self.calcColLayout()
        for colidx, colinfo in sorted(self.colLayout.items()):
            x, colwidth = colinfo
            col = self.columns[colidx]
            if x < windowWidth:  # only draw inside window
                # choose attribute to highlight column header
                if colidx == self.cursorColIndex:  # cursor is at this column
                    hdrattr = colors[options.c_CurHdr]
                elif colidx < self.nKeys:
                    hdrattr = colors[options.c_KeyCols]
                else:
                    hdrattr = colors[options.c_Header]

                y = 0
                vd.clipdraw(y, x, col.name or defaultColNames[colidx], hdrattr, colwidth)
                if x+colwidth+len(sepchars) <= windowWidth:
                    scr.addstr(y, x+colwidth, sepchars, colors[options.c_ColumnSep])

                y += 1
                for rowidx in range(0, windowHeight-2):
                    if self.topRowIndex + rowidx >= len(self.rows):
                        break

                    self.rowLayout[self.topRowIndex+rowidx] = y

                    row = self.rows[self.topRowIndex + rowidx]

                    if self.topRowIndex + rowidx == self.cursorRowIndex:  # cursor at this row
                        attr = colors[options.c_CurRow]
                    elif colidx < self.nKeys:
                        attr = colors[options.c_KeyCols]
                    else:
                        attr = colors[options.c_default]

                    if self.isSelected(row):
                        attr |= colors[options.c_SelectedRow]

                    if x+colwidth+len(sepchars) <= windowWidth:
                       scr.addstr(y, x+colwidth, sepchars, attr or colors[options.c_ColumnSep])

                    if colidx == self.cursorColIndex:  # cursor is at this column
                        attr |= colors[options.c_CurCol]

                    cellval = self.columns[colidx].getDisplayValue(row)
                    vd.clipdraw(y, x, cellval, attr, colwidth)
                    y += 1

    def editCell(self, colnum=None):
        if colnum is None:
            colnum = self.cursorColIndex
        x, w = self.colLayout[colnum]
        y = self.rowLayout[self.cursorRowIndex]
        currentValue = self.cellValue(self.cursorRowIndex, colnum)
        r = editText(y, x, w, value=currentValue, fillchar=options.EditPadChar)
        if r is None:
            vd.status('nothing changed')
            return currentValue
        if self.cursorCol.type:
            return self.cursorCol.type(r)  # convert input to column type
        else:
            return r  # presume string is fine

    def CommandHelp(self):
        vs = VSheet(self.name + '_help', self)
        vs.rows = list(self.commands.items())
        vs.columns = [VColumn('key', lambda r: keyname(r[0])),
                VColumn('action', lambda r: r[1][1]),
                VColumn('global_action', lambda r: global_commands[r[0]][1])]
        return vs

# end VSheet class

class VisiData:
    def __init__(self):
        self.sheets = VSheetSheets()
        self.statusHistory = []
        self._status = []
        self.status(initialStatus)
        self.lastErrors = []

    def status(self, s):
        strs = str(s)
        self._status.append(strs)
        self.statusHistory.insert(0, strs)
        del self.statusHistory[100:]  # keep most recent 100 only
        return s

    def exceptionCaught(self, status=True):
        import traceback
        self.lastErrors.append(traceback.format_exc().strip())
        self.lastErrors = self.lastErrors[-10:]  # keep most recent
        if status:
            self.status(self.lastErrors[-1].splitlines()[-1])
        if options.debug:
            raise

    def run(self):
        global sheet
        global windowHeight, windowWidth
        windowHeight, windowWidth = scr.getmaxyx()

        command_overrides = None
        while True:
            if not self.sheets:
                # if no more sheets, exit
                return

            sheet = self.sheets[0]
            if sheet.nRows == 0:
                self.status('no rows')

            try:
                sheet.draw()
            except Exception as e:
                self.exceptionCaught()

            # draw status on last line
            attr = colors[options.c_StatusLine]
            statusstr = options.SheetNameFmt % sheet.name + options.StatusSep.join(self._status)
            self.clipdraw(windowHeight-1, 0, statusstr, attr, windowWidth)
            self._status = []

            scr.move(windowHeight-1, windowWidth-2)
            curses.doupdate()

            ch = scr.getch()
            if ch == curses.KEY_RESIZE:
                windowHeight, windowWidth = scr.getmaxyx()
            elif ch == curses.KEY_MOUSE:
                try:
                    devid, x, y, z, bstate = curses.getmouse()
                    sheet.cursorRowIndex = sheet.topRowIndex+y-1
                except Exception:
                    self.exceptionCaught()
            elif (command_overrides and ch in command_overrides) or (not command_overrides and ch in sheet.commands):
                cmdstr, helpstr = command_overrides and command_overrides.get(ch) or sheet.commands.get(ch)
                try:
                    exec(cmdstr)

                    command_overrides = None
                except ChangeCommandSet as e:
                    # prefixes raise ChangeCommandSet exception instead
                    command_overrides = e.commands
                    self.status(e.mode)
                except VEscape as e:  # user aborted
                    self.status(str(e))
                except Exception:
                    command_overrides = None
                    self.exceptionCaught()
                    self.status(cmdstr)
            else:
                command_overrides = None
                self.status('no command for key "%s" (%d) ' % (keyname(ch), ch))

            sheet.checkCursor()

    def push(self, vs):
        if vs:
            if vs in self.sheets:
                self.sheets.remove(vs)
            self.sheets.insert(0, vs)
            return vs

    def clipdraw(self, y, x, s, attr=curses.A_NORMAL, w=None):
        s = s.replace('\n', '\\n')
        try:
            if w is None:
                w = windowWidth-1
            w = min(w, windowWidth-x-1)
            if w == 0:  # no room anyway
                return

            # convert to string just before drawing
            s = str(s)
            if len(s) > w:
                scr.addstr(y, x, s[:w-1] + options.Ellipsis, attr)
            else:
                scr.addstr(y, x, s, attr)
                if len(s) < w:
                    scr.addstr(y, x+len(s), options.ColumnFiller*(w-len(s)), attr)
        except Exception as e:
            self.status('clipdraw error: y=%s x=%s len(s)=%s w=%s' % (y, x, len(s), w))
            self.exceptionCaught()
# end VisiData class

## columns
class VColumn:
    def __init__(self, name, func=lambda r: r, width=None):
        self.name = name
        self.func = func
        self.width = width
        self.type = None
        self.expr = None  # Python string expression if computed column

    def getValue(self, row):
        try:
            return self.func(row)
        except Exception as e:
            vd.exceptionCaught(status=False)
            return options.FunctionError

    def getDisplayValue(self, row):
        cellval = self.getValue(row)
        if cellval is None:
            cellval = options.VisibleNone
        if isinstance(cellval, bytes):
            cellval = cellval.decode(options.encoding)
        return str(cellval).strip()

    def setValue(self, row, value):
        if options.readonly:
            vd.status('readonly mode')
        elif hasattr(self.func, 'setter'):
            self.func.setter(row, value)
        else:
            vd.status('column cannot be changed')


### common column setups and helpers
def setter(r, k, v):  # needed for use in lambda
    r[k] = v

def lambda_colname(colname):
    func = lambda r: r[colname]
    func.setter = lambda r,v,b=colname: setter(r,b,v)
    return func

def lambda_col(b):
    func = lambda r: r[b]
    func.setter = lambda r,v,b=b: setter(r,b,v)
    return func

def lambda_getattr(b):
    func = lambda r: getattr(r,b)
    func.setter = lambda r,v,b=b: setattr(r,b,v)
    return func

def lambda_subrow_wrap(func, subrowidx):
    'wraps original func to be func(r[subrowidx])'
    subrow_func = lambda r,i=subrowidx,f=func: r[i] and f(r[i]) or None
    subrow_func.setter = lambda r,v,i=subrowidx,f=func: r[i] and f.setter(r[i], v) or None
    return subrow_func

def ColumnExpr(sheet, expr):
    if expr:
        vc = VColumn(expr)
        vc.expr = expr
        vc.func = lambda r,col=vc,sheet=sheet: eval(col.expr, dict((c.name, c.getValue(r)) for c in sheet.columns if c is not col))
        return vc

def getPublicAttrs(obj):
    return [k for k in dir(obj) if not k.startswith('_') and not callable(getattr(obj, k))]

def PyobjColumns(exampleRow):
    'columns for each public attribute on an object'
    return [VColumn(k, lambda_getattr(k)) for k in getPublicAttrs(exampleRow)]

def ArrayColumns(n):
    'columns that display r[0]..r[n]'
    return [VColumn('', lambda_col(colnum)) for colnum in range(n)]

def ArrayNamedColumns(columns):
    'columnstr is a string of n column names (separated by spaces), mapping to r[0]..r[n]'
    return [VColumn(colname, lambda_col(i)) for i, colname in enumerate(columns)]

def AttrColumns(colnames):
    'colnames is list of attribute names'
    return [VColumn(name, lambda_getattr(name)) for name in colnames]

### sheet layouts
#### generic list/dict/object browsing
def pushPyObjSheet(name, pyobj, src=None):
    if isinstance(pyobj, list):
        return vd.push(VSheetList(name, pyobj, src=src))
    elif isinstance(pyobj, dict):
        return vd.push(VSheetDict(name, pyobj))
    elif isinstance(pyobj, object):
        return vd.push(VSheetObject(name, pyobj))
    else:
        vd.status('unknown type ' + type(pyobj))

class VSheetList(VSheet):
    def __init__(self, name, obj, columns=None, src=None):
        'columns is a list of strings naming attributes on the objects within the obj'
        super().__init__(name, src or obj)
        assert isinstance(obj, list)
        self.rows = obj
        if columns:
            self.columns = AttrColumns(columns)
        elif isinstance(obj[0], dict):  # list of dict
            self.columns = [VColumn(k, lambda_colname(k)) for k in obj[0].keys()]
        else:
            self.columns = [VColumn(name)]
        self.command(ENTER, 'pushPyObjSheet("%s[%s]" % (sheet.name, sheet.cursorRowIndex), sheet.cursorRow).cursorRowIndex = sheet.cursorColIndex', 'dive into this row')

class VSheetDict(VSheet):
    def __init__(self, name, mapping):
        super().__init__(name, mapping)
        self.rows = sorted(list(list(x) for x in mapping.items()))
        self.columns = [
            VColumn('key', lambda_col(0)),
            VColumn('value', lambda_col(1)) ]
        self.command(ENTER, 'if sheet.cursorColIndex == 1: pushPyObjSheet(sheet.name + options.SubsheetSep + sheet.cursorRow[0], sheet.cursorRow[1])', 'dive into this value')
        self.command(ord('e'), 'sheet.source[sheet.cursorRow[0]] = sheet.cursorRow[1] = sheet.editCell(1)', 'edit this value')

class VSheetObject(VSheet):
    def __init__(self, name, obj):
        super().__init__(name, obj)
        self.command(ENTER, 'pushPyObjSheet(sheet.name + options.SubsheetSep + sheet.cursorRow[0], sheet.cursorRow[1])', 'dive into this value')
        self.command(ord('e'), 'setattr(sheet.source, sheet.cursorRow[0], sheet.editCell(1)); sheet.reload()', 'edit this value')
        self.reload()

    def reload(self):
        valfunc = lambda_col(1)
        valfunc.setter = lambda r,v,obj=self.source: setattr(obj, r[0], v)
        self.columns = [
            VColumn(type(self.source).__name__ + '_attr', lambda_col(0)),
            VColumn('value', valfunc) ]
        self.rows = [(k, getattr(self.source, k)) for k in getPublicAttrs(self.source)]

#### specialized meta sheets
class VSheetSheets(VSheet, list):
    def __init__(self):
        VSheet.__init__(self, 'sheets', 'VisiData.sheets')
        self.command(ENTER,    'moveListItem(sheet, sheet.cursorRowIndex, 0); vd.sheets.pop(1)', 'go to this sheet')
        self.command(ord('&'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="&")', 'inner join')
        self.command(ord('+'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="+")', 'outer join')
        self.command(ord('*'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="*")', 'full join')
        self.command(ord('~'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="~")', 'diff join')
        self.reload()

    def reload(self):
        self.columns = AttrColumns('name nRows nCols cursorValue keyColNames source'.split())
        self.rows = self

class VSheetColumns(VSheet):
    def __init__(self, srcsheet):
        super().__init__(srcsheet.name + '_columns', srcsheet)
        self.command(ord('#'), 'sheet.source.convertType(sheet.cursorRow, int); sheet.moveCursorDown(+1)', 'convert column to int')
        self.command(ord('$'), 'sheet.source.convertType(sheet.cursorRow, str); sheet.moveCursorDown(+1)', 'convert column to str')
        self.command(ord('%'), 'sheet.source.convertType(sheet.cursorRow, float); sheet.moveCursorDown(+1)', 'convert column to float')
#        self.command(ord('@'), 'sheet.source.convertType(sheet.cursorRow, datetime); sheet.moveCursorDown(+1)', 'convert to datetime')
        self.reload()

    def reload(self):
        self.rows = self.source.columns
        self.columns = [
            VColumn('column', lambda_getattr('name')),
            VColumn('width',  lambda_getattr('width')),
            VColumn('type',   lambda_getattr('type')),
            VColumn('expr',   lambda_getattr('expr')),
            VColumn('value', lambda c,sheet=self: c.getValue(sheet.cursorRow)),
        ]
        if options.ColumnStats:
            self.columns.extend([
                VColumn('mode',   lambda c: statistics.mode(sheet.columnValues(c))),
                VColumn('min',    lambda c: min(sheet.columnValues(c))),
                VColumn('median', lambda c: statistics.median(sheet.columnValues(c))),
                VColumn('mean',   lambda c: statistics.mean(sheet.columnValues(c))),
                VColumn('max',    lambda c: max(sheet.columnValues(c))),
                VColumn('stddev', lambda c: statistics.stdev(sheet.columnValues(c))),
            ])

#### slicing and dicing
class VSheetJoin(VSheet):
    def __init__(self, sheets, jointype='&'):
        super().__init__(jointype.join(vs.name for vs in sheets))
        self.source = sheets
        self.jointype = jointype
        self.reload()

    def reload(self):
        sheets = self.source
        # first element in joined row is the tuple of keys
        self.columns = [VColumn(sheets[0].columns[colnum].name, lambda_subrow_wrap(lambda_col(colnum), 0)) for colnum in range(sheets[0].nKeys)]
        self.nKeys = sheets[0].nKeys

        rowsBySheetKey = {}
        rowsByKey = {}

        for vs in sheets:
            rowsBySheetKey[vs] = {}
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                rowsBySheetKey[vs][key] = r

        for sheetnum, vs in enumerate(sheets):
            # subsequent elements are the rows from each source, in order of the source sheets
            self.columns.extend(VColumn(c.name, lambda_subrow_wrap(c.func, sheetnum+1)) for c in vs.columns[vs.nKeys:])
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                if key not in rowsByKey:
                    rowsByKey[key] = [key] + [rowsBySheetKey[vs2].get(key) for vs2 in sheets]

        if self.jointype == '&':  # inner join  (only rows with matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if all(combinedRow))
        elif self.jointype == '+':  # outer join (all rows from first sheet)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items())
        elif self.jointype == '*':  # full join (keep all rows from all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items())
        elif self.jointype == '~':  # diff join (only rows without matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if not all(combinedRow))


class VSheetFreqTable(VSheet):
    def __init__(self, sheet, col):
        fqcolname = '%s_%s_freq' % (sheet.name, col.name)
        super().__init__(fqcolname, sheet)

        values = collections.defaultdict(list)
        for r in sheet.rows:
            values[str(col.getValue(r))].append(r)

        self.rows = sorted(values.items(), key=lambda r: len(r[1]), reverse=True)  # sort by num reverse
        self.largest = len(self.rows[0][1])+1

        self.columns = [
            VColumn(col.name, lambda_col(0)),
            VColumn('num', lambda r: len(r[1])),
            VColumn('histogram', lambda r,s=self: options.HistogramChar*int(len(r[1])*80/s.largest), width=80)
        ]
        self.command(ord(' '), 'sheet.source.toggle(sheet.cursorRow[1])', 'toggle these entries')
        self.command(ord('s'), 'sheet.source.select(sheet.cursorRow[1])', 'select these entries')
        self.command(ord('u'), 'sheet.source.unselect(sheet.cursorRow[1])', 'unselect these entries')


### input formats and helpers

sourceCache = {}

class Path:
    def __init__(self, fqpn):
        self.fqpn = fqpn
        self.name = os.path.split(fqpn)[-1]
        self.suffix = os.path.splitext(self.name)[1][1:]
    def read_text(self):
        return open(self.fqpn, encoding=options.encoding, errors=options.encoding_errors).read()
    def is_dir(self):
        return os.path.isdir(self.fqpn)
    def iterdir(self):
        return [Path(os.path.join(self.fqpn, f)) for f in os.listdir(self.fqpn)]
    def stat(self):
        return os.stat(self.fqpn)
    def __str__(self):
        return self.fqpn

def getTextContents(p):
    if not p in sourceCache:
        sourceCache[p] = p.read_text()
    return sourceCache[p]

def openFileOrUrl(p):
    if isinstance(p, Path):
        if p.is_dir():
            vs = VSheetDirectory(p)
        else:
            vs = globals().get('open_' + p.suffix, open_txt)(p)
    elif '://' in p:
        vs = openUrl(p)
    else:
        return openFileOrUrl(Path(p))
    return vd.push(vs)

class VSheetText(VSheet):
    def __init__(self, name, content, src=None):
        super().__init__(name, src)
        self.columns = [VColumn(name)]
        if isinstance(content, list):
            self.rows = content
        elif isinstance(content, str):
            self.rows = content.split('\n')
        else:
            raise VException('unknown text type ' + str(type(content)))

class VSheetDirectory(VSheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        self.columns = [VColumn('filename', lambda r: r[0].name),
                        VColumn('type', lambda r: r[0].is_dir() and '/' or r[0].suffix),
                        VColumn('size', lambda r: r[1].st_size),
                        VColumn('mtime', lambda r: datestr(r[1].st_mtime))]
        self.command(ENTER, 'openFileOrUrl(sheet.cursorRow[0])')  # path, filename
        self.reload()

    def reload(self):
        self.rows = [(p, p.stat()) for p in self.source.iterdir() if not p.name.startswith('.')]

def open_txt(p):
    contents = getTextContents(p)
    if '\t' in contents[:32]:
        return open_tsv(p)  # TSV often have .txt extension
    return VSheetText(p.name, contents, p)

class open_csv(VSheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        contents = getTextContents(p)

        if options.csv_dialect == 'sniff':
            headers = contents[:1024]
            dialect = csv.Sniffer().sniff(headers)
            vd.status('sniffed csv_dialect as %s' % dialect)
        else:
            dialect = options.csv_dialect

        rdr = csv.reader(io.StringIO(contents, newline=''), dialect=dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        self.rows = [r for r in rdr]
        self.columns = PyobjColumns(self.rows[0]) or ArrayColumns(len(self.rows[0]))

class open_tsv(VSheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        lines = getTextContents(p).splitlines()

        if options.csv_header:
            self.columns = ArrayNamedColumns(lines[0].split('\t'))
            lines = lines[1:]
        else:
            self.columns = ArrayColumns(len(lines[0].split('\t')))

        self.rows = [L.split('\t') for L in lines]  # [rownum] -> [ field, ... ]

def open_json(p):
    import json
    pushPyObjSheet(p.name, json.loads(getTextContents(p)))

#### .xlsx
class open_xlsx(VSheet):
    def __init__(self, path):
        super().__init__(path.name, path)
        import openpyxl
        self.workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        self.rows = list(self.workbook.sheetnames)
        self.columns = [VColumn('name')]
        self.command(ENTER, 'vd.push(sheet.getSheet(sheet.cursorRow))', 'open this sheet')

    def getSheet(self, sheetname):
        'create actual VSheet from xlsx sheet'
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        vs = VSheet('%s%s%s' % (self.source, options.SubsheetSep, sheetname), worksheet)
        vs.columns = ArrayColumns(worksheet.max_column)
        vs.rows = [ [cell.value for cell in row] for row in worksheet.iter_rows()]
        return vs

#### .hdf5
class VSheetH5Obj(VSheet):
    def __init__(self, name, hobj, src):
        super().__init__(name, src)
        self.hobj = hobj
        self.reload()

    def reload(self):
        import h5py
        if isinstance(self.hobj, h5py.Group):
            self.rows = [ self.hobj[objname] for objname in self.hobj.keys() ]
            self.columns = [
                VColumn(self.hobj.name, lambda r: r.name.split('/')[-1]),
                VColumn('type', lambda r: type(r).__name__),
                VColumn('nItems', lambda r: len(r)),
            ]
            self.command(ENTER, 'vd.push(VSheetH5Obj(sheet.name+options.SubsheetSep+sheet.cursorRow.name, sheet.cursorRow, sheet.source))', 'dive into this object')
            self.command(ord('A'), 'vd.push(VSheetDict(sheet.cursorRow.name + "_attrs", sheet.cursorRow.attrs))', 'view/edit the metadata for this object')
        elif isinstance(self.hobj, h5py.Dataset):
            if len(self.hobj.shape) == 1:
                self.rows = self.hobj[:]  # copy
                self.columns = [VColumn(colname, lambda_colname(colname)) for colname in self.hobj.dtype.names]
            elif len(self.hobj.shape) == 2:  # matrix
                self.rows = self.hobj[:]  # copy
                self.columns = ArrayColumns(self.hobj.shape[1])
            else:
                vd.status('too many dimensions in shape %s' % str(self.hobj.shape))
        else:
            vd.status('unknown h5 object type %s' % type(self.hobj))

class open_hdf5(VSheetH5Obj):
    def __init__(self, p):
        import h5py
        super().__init__(p.name, h5py.File(str(p), 'r'), p)
open_h5 = open_hdf5

#### databases
class VSheetBlaze(VSheet):
    def __init__(self, name, data, src):
        super().__init__(name, src)
        self.columns = ArrayNamedColumns(data.fields)
        self.rows = list(data)

def openUrl(url):
    import blaze
    import datashape; datashape.coretypes._canonical_string_encodings.update({"utf8_unicode_ci": "U8"})
    fp = blaze.data(url)
    vs = VSheetList(url, [getattr(fp, tblname) for tblname in fp.fields], url)
    vs.command(ENTER, 'vd.push(VSheetBlaze(sheet.cursorRow.name, sheet.cursorRow, sheet))', 'dive into this table')
    return vs

### Sheet savers

def saveSheet(sheet, fn):
    basename, ext = os.path.splitext(fn)
    funcname = 'save_' + ext[1:]
    globals().get(funcname, save_tsv)(sheet, fn)
    vd.status('saved to ' + fn)

def save_tsv(sheet, fn):
    with open(fn, 'w', encoding=options.encoding, errors=options.encoding_errors) as fp:
        colhdr = '\t'.join(col.name for col in sheet.columns) + '\n'
        if colhdr.strip():  # is anything but whitespace
            fp.write(colhdr)
        for r in sheet.rows:
            fp.write('\t'.join(col.getDisplayValue(r) for col in sheet.columns) + '\n')

def save_csv(sheet, fn):
    with open(fn, 'w', newline='', encoding=options.encoding, errors=options.encoding_errors) as fp:
        cw = csv.writer(fp, dialect=options.csv_dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        colnames = [col.name for col in sheet.columns]
        if ''.join(colnames):
            cw.writerow(colnames)
        for r in sheet.rows:
            cw.writerow([col.getDisplayValue(r) for col in sheet.columns])

### curses, options, init
def editText(y, x, w, prompt='', value='', fillchar=' '):
    def splice(v, i, s):  # splices s into the string v at i (v[i] = s[0])
        return v if i < 0 else v[:i] + s + v[i:]
    def clean(s):
        return ''.join(c if c.isprintable() else options.Unprintable for c in s)
    def delchar(s, i, remove=1):
        return s[:i] + s[i+remove:]

    if prompt:
        scr.addstr(y, x, prompt)
        x += len(prompt)
        w -= len(prompt)

    v = value   # value under edit
    i = 0       # index into v
    while True:
        dispval = clean(v)
        dispi = i
        if len(v) < w:
            dispval += fillchar*(w-len(v))
        elif i >= w:
            dispi = w-1
            dispval = dispval[i-w:]

        scr.addstr(y, x, dispval, colors[options.c_EditCell])
        scr.move(y, x+dispi)
        ch = scr.getch()
        if ch == ctrl('a') or ch == curses.KEY_HOME:        i = 0
        elif ch == ctrl('b') or ch == curses.KEY_LEFT:      i -= 1
        elif ch == ctrl('c') or ch == ESC:                  raise VEscape(keyname(ch))
        elif ch == ctrl('d') or ch == curses.KEY_DC:        v = delchar(v, i)
        elif ch == ctrl('e') or ch == curses.KEY_END:       i = len(v)
        elif ch == ctrl('f') or ch == curses.KEY_RIGHT:     i += 1
        elif ch == ctrl('h') or ch == curses.KEY_BACKSPACE: i -= 1 if i > 0 else 0; v = delchar(v, i)
        elif ch == ctrl('j') or ch == ENTER:                break
        elif ch == ctrl('k'):                               v = v[:i]
        elif ch == ctrl('r'):                               v = value
        elif ch == ctrl('t'):                               v = delchar(splice(v, i-2, v[i-1]), i)
        elif ch == ctrl('u'):                               v = v[i:]; i = 0
        elif ch == ctrl('v'):                               v = splice(v, i, chr(scr.getch())); i+= 1
        else:
            v = splice(v, i, chr(ch))
            i += 1

        if i < 0: i = 0
        if i > len(v): i = len(v)

    vd.status('"%s"' % v)
    return v

def inputLine(prompt=''):
    'move to the bottom of the screen and get a line of input from the user'
    return editText(windowHeight-1, 0, windowWidth-1, prompt, fillchar=' ')


nextColorPair = 1
def setupcolors(stdscr, f, *args):
    def makeColor(fg, bg):
        global nextColorPair
        if curses.has_colors():
            curses.init_pair(nextColorPair, fg, bg)
            c = curses.color_pair(nextColorPair)
            nextColorPair += 1
        else:
            c = curses.A_NORMAL

        return c

    curses.meta(1)  # allow "8-bit chars"

    colors['red'] = curses.A_BOLD | makeColor(curses.COLOR_RED, curses.COLOR_BLACK)
    colors['blue'] = curses.A_BOLD | makeColor(curses.COLOR_BLUE, curses.COLOR_BLACK)
    colors['green'] = curses.A_BOLD | makeColor(curses.COLOR_GREEN, curses.COLOR_BLACK)
    colors['brown'] = makeColor(curses.COLOR_YELLOW, curses.COLOR_BLACK)
    colors['yellow'] = curses.A_BOLD | colors['brown']
    colors['cyan'] = makeColor(curses.COLOR_CYAN, curses.COLOR_BLACK)
    colors['magenta'] = makeColor(curses.COLOR_MAGENTA, curses.COLOR_BLACK)

    colors['red_bg'] = makeColor(curses.COLOR_WHITE, curses.COLOR_RED)
    colors['blue_bg'] = makeColor(curses.COLOR_WHITE, curses.COLOR_BLUE)
    colors['green_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_GREEN)
    colors['brown_bg'] = colors['yellow_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_YELLOW)
    colors['cyan_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_CYAN)
    colors['magenta_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_MAGENTA)

    return f(stdscr, *args)


def terminal_main():
    'Parse arguments and initialize VisiData instance'
    import argparse

    global vd
    parser = argparse.ArgumentParser(description=__doc__)

    parser.add_argument('inputs', nargs='*', help='initial sheets')
    parser.add_argument('-d', '--debug', dest='debug', action='store_true', default=False, help='abort on exception')
    parser.add_argument('-r', '--readonly', dest='readonly', action='store_true', default=False, help='editing disabled')
    args = parser.parse_args()

    options.debug = args.debug
    options.readonly = args.readonly

    vd = VisiData()
    inputs = args.inputs or ['.']

    for arg in inputs:
        openFileOrUrl(arg)

    os.putenv('ESCDELAY', '25')  # reduce ESC timeout to 25ms. http://en.chys.info/2009/09/esdelay-ncurses/
    ret = wrapper(curses_main)
    if ret:
        print(ret)


def curses_main(_scr):
    global scr
    scr = _scr

    # get control keys instead of signals
    curses.raw()

    # enable mouse events
#    curses.mousemask(curses.ALL_MOUSE_EVENTS)

    try:
        return vd.run()
    except Exception as e:
        if options.debug:
            raise
        return 'Exception: ' + str(e)


def wrapper(f, *args):
    import curses
    return curses.wrapper(setupcolors, f, *args)


if __name__ == '__main__':
    terminal_main()

import itertools
import copy
import datetime
import re
from colorsys import rgb_to_hls, hls_to_rgb

from visidata import VisiData, vd, Sheet, Column, Progress, IndexSheet, ColumnAttr, SequenceSheet, AttrDict, AttrColumn
from visidata import CellColorizer, getattrdeep, rgb_to_attr
from visidata.type_date import date


vd.option('xlsx_meta_columns', False, 'include columns for cell objects, font colors, and fill colors', replay=True)

@VisiData.api
def open_xls(vd, p):
    return XlsIndexSheet(p.base_stem, source=p)

@VisiData.api
def open_xlsx(vd, p):
    return XlsxIndexSheet(p.base_stem, source=p)

class XlsxIndexSheet(IndexSheet):
    'Load XLSX file (in Excel Open XML format).'
    rowtype = 'sheets'  # rowdef: xlsxSheet
    columns = [
        Column('sheet', getter=lambda col,row: row.source.title),  # xlsx sheet title
        ColumnAttr('name', width=0),  # visidata Sheet name
        ColumnAttr('nRows', type=int),
        ColumnAttr('nCols', type=int),
        Column('active', getter=lambda col,row: row.source is col.sheet.workbook.active),
    ]
    nKeys = 1

    def iterload(self):
        openpyxl = vd.importExternal('openpyxl')
        self.workbook = openpyxl.load_workbook(str(self.source), data_only=True, read_only=True)
        for sheetname in self.workbook.sheetnames:
            src = self.workbook[sheetname]
            yield XlsxSheet(self.name, sheetname, source=src, workbook=self.workbook)


class XlsxSheet(SequenceSheet):
    # rowdef: AttrDict of column_letter to cell
    colorizers = [
        CellColorizer(5, None, lambda s,c,r,v: c and r and s.colorize_xlsx_cell(c,r))
    ]
    def setCols(self, headerrows):
        vd.importExternal('openpyxl')
        from openpyxl.utils.cell import get_column_letter
        self.columns = []
        self._rowtype = AttrDict

        if not headerrows:
            return

        headers = [[cell.value for cell in row.values()] for row in headerrows]
        column_letters = [
                x.column_letter if 'column_letter' in dir(x)
                else get_column_letter(i+1)
                for i, x in enumerate(headerrows[0].values())]

        for i, colnamelines in enumerate(itertools.zip_longest(*headers, fillvalue='')):
            colnamelines = ['' if c is None else c for c in colnamelines]
            column_name = ''.join(map(str, colnamelines))
            self.addColumn(AttrColumn(column_name, column_letters[i] + '.value', column_letter=column_letters[i]))
            self.addXlsxMetaColumns(column_letters[i], column_name)

    def addRow(self, row, index=None):
        Sheet.addRow(self, row, index=index)  # skip SequenceSheet
        for column_letter, v in list(row.items())[len(self.columns):len(row)]:  # no-op if already done
            self.addColumn(AttrColumn('', column_letter + '.value'))
            self.addXlsxMetaColumns(column_letter, column_letter)

    def iterload(self):
        vd.importExternal('openpyxl')
        from openpyxl.utils.cell import get_column_letter
        worksheet = self.source
        for row in Progress(worksheet.iter_rows(), total=worksheet.max_row or 0):
            yield AttrDict({get_column_letter(i+1): cell for i, cell in enumerate(row)})

    def addXlsxMetaColumns(self, column_letter, column_name):
        if self.options.xlsx_meta_columns:
            self.addColumn(
                    AttrColumn(column_name + '_cellPyObj', column_letter, column_letter=column_letter))
            self.addColumn(
                    AttrColumn(column_name + '_fontColor',
                        column_letter + '.font.color.value', column_letter=column_letter))
            self.addColumn(
                    AttrColumn(column_name + '_fillColor', column_letter +
                        '.fill.start_color.value', column_letter=column_letter))
            self.addColumn(Column(column_name + '_colorizer', width=0,
                           column_letter=column_letter,
                           getter=lambda c,r: c.sheet.colorize_xlsx_cell(c,r)))

    def paste_after(self, rowidx):
        to_paste = list(copy.copy(r) for r in reversed(vd.memory.cliprows))
        self.addRows(to_paste, index=rowidx)


class XlsIndexSheet(IndexSheet):
    'Load XLS file (in Excel format).'
    rowtype = 'sheets'  # rowdef: xlsSheet
    columns = [
        Column('sheet', getter=lambda col,row: row.source.name),  # xls sheet name
        ColumnAttr('name', width=0),  # visidata sheet name
        ColumnAttr('nRows', type=int),
        ColumnAttr('nCols', type=int),
    ]
    nKeys = 1
    def iterload(self):
        xlrd = vd.importExternal('xlrd')
        self.workbook = xlrd.open_workbook(str(self.source))
        for sheetname in self.workbook.sheet_names():
            yield XlsSheet(self.name, sheetname, source=self.workbook.sheet_by_name(sheetname))


class XlsSheet(SequenceSheet):
    def iterload(self):
        worksheet = self.source
        for rownum in Progress(range(worksheet.nrows)):
            yield list(worksheet.cell(rownum, colnum).value for colnum in range(worksheet.ncols))


@Sheet.property
def xls_name(vs):
    name = vs.names[-1]
    if vs.options.clean_names:
        cleaned_name = ''.join('_' if ch in ':[]*?/\\' else ch for ch in vs.name) #1122
        name = cleaned_name[:31] #1122  #594
        name = name.strip('_')

    return name


@VisiData.api
def save_xlsx(vd, p, *sheets):
    openpyxl = vd.importExternal('openpyxl')

    wb = openpyxl.Workbook()
    wb.remove_sheet(wb['Sheet'])

    def _convert_save_row(dispvals:dict, replace_illegal=False) -> list:
        row = []
        for col, v in dispvals.items():
            if v is None:
                v = ""
            elif col.type == date:
                v = datetime.datetime.fromtimestamp(int(v.timestamp()))
            elif not vd.isNumeric(col):
                v = str(v)
                if replace_illegal:
                    v = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, ' ', v)

            row.append(v)
        return row

    for vs in sheets:
        if vs.xls_name != vs.names[-1]:
            vd.warning(f'saving {vs.name} as {vs.xls_name}')
        ws = wb.create_sheet(title=vs.xls_name)

        headers = [col.name for col in vs.visibleCols]
        ws.append(headers)

        for dispvals in vs.iterdispvals(format=False):
            row = _convert_save_row(dispvals)
            try:
                ws.append(row)
            except openpyxl.utils.exceptions.IllegalCharacterError as e:
                row = _convert_save_row(dispvals, replace_illegal=True)  #1402
                ws.append(row)

    wb.active = ws

    wb.save(filename=p)


@VisiData.api
def save_xls(vd, p, *sheets):
    xlwt = vd.importExternal('xlwt')

    wb = xlwt.Workbook()

    for vs in sheets:
        if vs.xls_name != vs.name:
            vd.warning(f'saving {vs.name} as {vs.xls_name}')
        ws1 = wb.add_sheet(vs.xls_name)
        for col_i, col in enumerate(vs.visibleCols):
            ws1.write(0, col_i, col.name)

        for r_i, dispvals in enumerate(vs.iterdispvals(format=True)):
            r_i += 1
            for c_i, v in enumerate(dispvals.values()):
                ws1.write(r_i, c_i, v)

    wb.save(p)


# from https://stackoverflow.com/a/65426130

RGBMAX = 255
HLSMAX = 240

@XlsxSheet.api
def colorize_xlsx_cell(sheet, col, row):
    fg = getattrdeep(row, col.column_letter+'.font.color', None)
    bg = getattrdeep(row, col.column_letter+'.fill.start_color', None)
    fg = sheet.xlsx_color_to_xterm256(fg)
    bg = sheet.xlsx_color_to_xterm256(bg)

    if bg == '-1' or fg == '-1':
        fg, bg = '-1', '-1'

    return f'{fg} on {bg}'

@XlsxSheet.api
def xlsx_color_to_xterm256(sheet, color) -> str:
    if not color:
        return ''

    if color.type == 'rgb':
        s = color.value
        if isinstance(s, int):
            return str(s)

        a,r,g,b = s[0:2], s[2:4], s[4:6], s[6:8]
        return rgb_to_attr(int(r, 16), int(g, 16), int(b, 16), int(a, 16))

    if color.type == 'theme':
        return sheet.theme_and_tint_to_rgb(color.value, color.tint)
    else:
        return str(color.value)

@XlsxSheet.api
def theme_and_tint_to_rgb(sheet, theme, tint) -> str:
    """Given a workbook, a theme number and a tint return a xterm256 color number"""
    rgb = sheet.theme_colors[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    r, g, b = ms_hls_to_rgb(h, tint_luminance(tint, l), s)

    return rgb_to_attr(r*256, g*256, b*256)

@XlsxSheet.lazy_property
def theme_colors(sheet):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(sheet.workbook.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]

    theme_colors = []

    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        for i in list(accent): # walk all child nodes, rather than assuming [0]
            if 'window' in i.attrib['val']:
                theme_colors.append(i.attrib['lastClr'])
            else:
                theme_colors.append(i.attrib['val'])

    return theme_colors

def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()

def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))

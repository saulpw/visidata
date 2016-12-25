#!/usr/bin/env python3
#
# VisiData: a curses interface for exploring and arranging tabular data
#
# Copyright (C) 2016 Saul Pwanson
#

__version__ = '0.39'
__author__ = 'Saul Pwanson <vd@saul.pw>'
__license__ = 'GPLv3'
__status__ = 'Development'

import os
import os.path
from copy import copy
import io
import collections
import functools
import statistics
import curses

import re
import csv

class attrdict(object):
    def __init__(self, kwargs):
        self.__dict__ = kwargs

    @classmethod
    def kwargs(cls, **kwargs):
        return cls(kwargs)

base_options = collections.OrderedDict()
user_options = None
options = attrdict(base_options)


# minimalist 'any'
anytype = lambda r=None: str(r)
anytype.__name__ = ''

# some str wrappers for annotations on cell results
class WrongTypeStr(str):
    pass

class CalcErrorStr(str):
    pass


allPrefixes = 'g'

# VisiData singleton contains all sheets
@functools.lru_cache()
def vd():
    return VisiData()


def status(s):
    return vd().status(s)

def error(s):
    raise Exception(s)

def moveListItem(L, fromidx, toidx):
    r = L.pop(fromidx)
    L.insert(toidx, r)
    return toidx

def values(r, cols):
    assert cols, cols
    d = {}
    for c in cols:
        d[c.name] = c.getValue(r)
    return attrdict(d)

from visidata.tui import edit_text, Key, Shift, Ctrl, keyname, EscapeException, wrapper, colors, draw_clip
from visidata.date import date
from visidata.Path import Path
from .Sheet import Sheet, setup_sheet_commands
from .Column import Column, ColumnAttr, ColumnSourceAttr, ColumnItem, ColumnExpr, ColumnRegex
from .Column import AttrColumns, PyobjColumns, ArrayColumns, ArrayNamedColumns
from .Column import lambda_if_firstcol, lambda_subrow_wrap, lambda_colname


initialStatus = 'saul.pw/VisiData v' + __version__

sheet = None
windowWidth = None
windowHeight = None

def set_sheet(s):
    'for tests'
    global sheet
    sheet = s

def setup_options():
    def option(name, default, helpstr=''):
        if name not in base_options:
            base_options[name] = default
    def theme(name, default, helpstr=''):
        base_options[name] = default

    option('csv_dialect', 'excel', 'dialect passed to csv.reader')
    option('csv_delimiter', ',', 'delimiter passed to csv.reader')
    option('csv_quotechar', '"', 'quotechar passed to csv.reader')
    option('csv_header', '', 'parse first row of CSV as column names')

    option('debug', '', 'abort on error and display stacktrace')
    option('readonly', '', 'disable saving')
    option('userdir_prefix', '~/.', 'prefix for vd-options.tsv')

    option('encoding', 'utf-8', 'as passed to codecs.open')
    option('encoding_errors', 'surrogateescape', 'as passed to codecs.open')
    option('save_dir', '.', 'default output folder')

    option('ColumnStats', False, 'include mean/median/etc on Column sheet')

    option('SubsheetSep', '~')

    # display/color scheme
    theme('SheetNameFmt', '%s| ', 'status line prefix')
    theme('ch_VisibleNone', '',  'visible contents of a cell whose value was None')
    theme('ch_FunctionError', '¿', 'when computation fails due to exception')
    theme('ch_Histogram', '*')
    theme('ch_ColumnFiller', ' ', 'pad chars after column value')
    theme('ch_LeftMore', '<')
    theme('ch_RightMore', '>')
    theme('ch_ColumnSep', '|', 'chars between columns')
    theme('ch_Ellipsis', '…')
    theme('ch_StatusSep', ' | ')
    theme('ch_KeySep', '/')
    theme('ch_EditPadChar', '_')
    theme('ch_Newline', '\\n')
    theme('ch_Unprintable', '.')
    theme('ch_WrongType', '~')
    theme('ch_Error', '!')

    theme('c_default', 'normal')
    theme('c_Header', 'bold')
    theme('c_CurHdr', 'reverse')
    theme('c_CurRow', 'reverse')
    theme('c_CurCol', 'bold')
    theme('c_KeyCols', 'brown')
    theme('c_StatusLine', 'bold')
    theme('c_SelectedRow', 'green')
    theme('c_ColumnSep', 'blue')
    theme('c_EditCell', 'normal')
    theme('c_WrongType', 'magenta')
    theme('c_Error', 'red')
# end setup_sheet_commands

### VisiData core

def detectType(v):
    def tryType(T, v):
        try:
            v = T(v)
            return T
        except:
            return None

    return tryType(int, v) or tryType(float, v) or tryType(date, v) or str


class VisiData:
    # base_commands is a list that contains the commands defined by the visidata package (including some that are sheet-specific).
    # user_commands contains all user-specified overrides.  Any changes to base_commands are actually additions to user_commands.
    # They are used by the sheet's load_commands() to make its sheet.commands, and also serve as the sources for the gF1 SheetAppend.
    base_commands = None
    user_commands = None

    def __init__(self):
        self.sheets = []
        self.statusHistory = []
        self._status = []
        self.status(initialStatus)
        self.status('F1 opens command help')
        self.lastErrors = []

    def status(self, s):
        strs = str(s)
        self._status.append(strs)
        self.statusHistory.insert(0, strs)
        del self.statusHistory[100:]  # keep most recent 100 only
        return s

    def editText(self, y, x, w, **kwargs):
        v = edit_text(self.scr, y, x, w, **kwargs)
        self.status('"%s"' % v)
        return v

    def exceptionCaught(self, status=True):
        import traceback
        self.lastErrors.append(traceback.format_exc().strip())
        self.lastErrors = self.lastErrors[-10:]  # keep most recent
        if status:
            self.status(self.lastErrors[-1].splitlines()[-1])
        if options.debug:
            raise

    def drawLeftStatus(self, sheet):
        'draws sheet info on last line, including previous status messages, which are then cleared.'
        attr = colors[options.c_StatusLine]
        statusstr = options.SheetNameFmt % sheet.name + options.ch_StatusSep.join(self._status)
        draw_clip(self.scr, windowHeight-1, 0, statusstr, attr, windowWidth)
        self._status = []

    def drawRightStatus(self, rstatus):
        draw_clip(self.scr, windowHeight-1, windowWidth-len(rstatus)-2, rstatus, colors[options.c_StatusLine])

    def run(self, scr):
        global windowHeight, windowWidth, sheet
        windowHeight, windowWidth = scr.getmaxyx()
        self.scr = scr

        command_overrides = None
        prefixes = ''
        keystroke = ''
        while True:
            if not self.sheets:
                # if no more sheets, exit
                return

            sheet = self.sheets[0]
            if sheet.nRows == 0:
                self.status('no rows')

            try:
                sheet.draw(scr)
            except Exception as e:
                self.exceptionCaught()

            self.drawLeftStatus(sheet)

            self.drawRightStatus(prefixes+keystroke)  # visible for this getch

            curses.doupdate()
            ch = scr.getch()
            keystroke = keyname(ch)
            self.drawRightStatus(prefixes+keystroke)  # visible for commands that wait with getch
            if ch == curses.KEY_RESIZE:
                windowHeight, windowWidth = scr.getmaxyx()
            elif ch == curses.KEY_MOUSE:
                try:
                    devid, x, y, z, bstate = curses.getmouse()
                    sheet.cursorRowIndex = sheet.topRowIndex+y-1
                except Exception:
                    self.exceptionCaught()
            elif keystroke in allPrefixes:
                prefixes += keystroke
            else:
                try:
                    sheet.exec_command(globals(), prefixes, keystroke)
                except EscapeException as e:  # user aborted
                    self.status(keyname(e.args[0]))
                except Exception:
                    if options.debug:
                        raise
                    self.exceptionCaught()
#                    self.status(sheet.commands[prefixes+keystroke].execstr)
                prefixes = ''

            sheet.checkCursor()

    def replace(self, vs):
        'replace top sheet with the given sheet'
        self.sheets.pop(0)
        return vs.push(vs)

    def push(self, vs):
        if vs:
            if vs in self.sheets:
                self.sheets.remove(vs)
            elif not vs.rows:  # first time
                vs.reload_commands()
                vs.reload()
            self.sheets.insert(0, vs)
            return vs
# end VisiData class

### sheet layouts
#### generic list/dict/object browsing
def push_pyobj(name, pyobj, src=None):
    vs = load_pyobj(name, pyobj, src)
    if vs:
        return vd().push(vs)
    else:
        status('unknown type ' + type(pyobj))

def load_pyobj(name, pyobj, src=None):
    if isinstance(pyobj, list):
        return SheetList(name, pyobj, src=src)
    elif isinstance(pyobj, dict):
        return SheetDict(name, pyobj)
    elif isinstance(pyobj, object):
        return SheetObject(name, pyobj)
    else:
        status('unknown type ' + type(pyobj))

class SheetList(Sheet):
    def __init__(self, name, obj, columns=None, src=None):
        'columns is a list of strings naming attributes on the objects within the obj'
        super().__init__(name, src or obj)
        assert isinstance(obj, list)
        self.obj = obj
        self.colnames = columns

    def reload(self):
        self.rows = self.obj
        if self.colnames:
            self.columns = AttrColumns(self.colnames)
        elif isinstance(self.rows[0], dict):  # list of dict
            self.columns = [Column(k, type(self.rows[0][k]), lambda_colname(k)) for k in self.rows[0].keys()]
            self.nKeys = 1
        elif isinstance(self.rows[0], attrdict):  # list of attrdict
            self.columns = PyobjColumns(self.rows[0])
        else:
            self.columns = [Column(self.name)]

        self.command(Key.ENTER, 'push_pyobj("%s[%s]" % (name, cursorRowIndex), cursorRow).cursorRowIndex = cursorColIndex', 'dive into this row')

class SheetDict(Sheet):
    def __init__(self, name, mapping):
        super().__init__(name, mapping)
        self.rows = list(list(x) for x in mapping.items())
        self.columns = [ ColumnItem('key', 0), ColumnItem('value', 1) ]
        self.nKeys = 1
        self.command(Key.ENTER, 'if cursorColIndex == 1: push_pyobj(name + options.SubsheetSep + cursorRow[0], cursorRow[1])', 'dive into this value')
        self.command(Key('e'), 'source[cursorRow[0]] = cursorRow[1] = editCell(1)', 'edit this value')


class SheetObject(Sheet):
    def __init__(self, name, obj):
        super().__init__(name, obj)
        self.command(Key.ENTER, 'v = getattr(source, cursorRow); push_pyobj(name + options.SubsheetSep + cursorRow, v() if callable(v) else v)', 'dive into this value')
        self.command(Key('e'), 'setattr(source, cursorRow, editCell(1)); reload()', 'edit this value')

    def reload(self):
        super().reload()
        self.columns = [
            Column(type(self.source).__name__ + '_attr'),
            ColumnSourceAttr('value', self.source)
        ]
        self.nKeys = 1
        self.rows = dir(self.source)


#### specialized meta sheets

# commands are a list of attrdicts loaded in from to that will of rows like anything other sheet.  a dict cache would be more efficient for command lookup, but there are only ~100 commands to consider,
#    so a linear search may not be so bad, even if it does happen with every command.  and the expensive part of the search would be the regex match, which is
#    factored out when the individual sheet's commands list is composed.

#  so:
# - user_commands 
# - F1 pushes CommandsSheet with user_commands + sheet-specific-commands + base_commands, in that order.  sheet-specific can be from either user or base, but of course user overrides sheet and sheet overrides other ones from base.  non-matching sheet_regexes and overridden commands are filtered out.
# - Only the most recent entry is executed (with warning) if multiple matches by sheet regex and keystrokes.
# - gF1 pushes all_commands, which is an append of user_commands and base_commands.  Edits there are reflected back to user_commands.
# - any changes to these should be saved to ~/.vd-commands.tsv (but with prompt) with ^S or on sheet quit if options.autosave

class Sheets(SheetList):
    def __init__(self, src):
        super().__init__('sheets', vd().sheets, 'name nRows nCols cursorValue keyColNames source'.split())
        self.nKeys = 1

    def reload(self):
        super().reload()
        self.command(Key.ENTER,    'moveListItem(vd.sheets, cursorRowIndex, 0); vd.sheets.pop(1)', 'go to this sheet')
        self.command(Key('&'), 'vd.replace(SheetJoin(selectedRows, jointype="&"))', 'open inner join of selected sheets')
        self.command(Key('+'), 'vd.replace(SheetJoin(selectedRows, jointype="+"))', 'open outer join of selected sheets')
        self.command(Key('*'), 'vd.replace(SheetJoin(selectedRows, jointype="*"))', 'open full join of selected sheets')
        self.command(Key('~'), 'vd.replace(SheetJoin(selectedRows, jointype="~"))', 'open diff join of selected sheets')

class SheetColumns(Sheet):
    def __init__(self, srcsheet):
        super().__init__(srcsheet.name + '_columns', srcsheet)

        # on the Columns sheet, these affect the 'row' (column in the source sheet)
        self.command(Key('@'), 'cursorRow.type = date; cursorDown(+1)', 'set source column type to datetime')
        self.command(Key('#'), 'cursorRow.type = int; cursorDown(+1)', 'set source column type to integer')
        self.command(Key('$'), 'cursorRow.type = str; cursorDown(+1)', 'set source column type to string')
        self.command(Key('%'), 'cursorRow.type = float; cursorDown(+1)', 'set source column type to decimal numeric type')
        self.command(Key('~'), 'cursorRow.type = detectType(cursorRow.getValue(source.cursorRow)); cursorDown(+1)', 'autodetect type of source column using its data')
        self.command(Key('!'), 'source.toggleKeyColumn(cursorRowIndex)', 'toggle key column on source sheet')
        self.command(Key('-'), 'cursorRow.width = 0', 'hide column on source sheet')
        self.command(Key('_'), 'cursorRow.width = cursorRow.getMaxWidth(source.rows)', 'set source column width to max width of its rows')

    def reload(self):
        super().reload()
        self.rows = self.source.columns
        self.nKeys = 1
        self.columns = [
            ColumnAttr('name', str),
            ColumnAttr('width', str),
            Column('type',   str, lambda r: r.type.__name__),
            ColumnAttr('fmtstr', str),
            ColumnAttr('expr', str),
            Column('value',  anytype, lambda c,sheet=self.source: c.getValue(sheet.cursorRow)),
#            Column('nulls',  int, lambda c,sheet=sheet: c.nEmpty(sheet.rows)),

#            Column('uniques',  int, lambda c,sheet=sheet: len(set(c.values(sheet.rows))), width=0),
#            Column('mode',   anytype, lambda c: statistics.mode(c.values(sheet.rows)), width=0),
#            Column('min',    anytype, lambda c: min(c.values(sheet.rows)), width=0),
#            Column('median', anytype, lambda c: statistics.median(c.values(sheet.rows)), width=0),
#            Column('mean',   float, lambda c: statistics.mean(c.values(sheet.rows)), width=0),
#            Column('max',    anytype, lambda c: max(c.values(sheet.rows)), width=0),
#            Column('stddev', float, lambda c: statistics.stdev(c.values(sheet.rows)), width=0),
            ]

#### slicing and dicing
class SheetJoin(Sheet):
    def __init__(self, sheets, jointype='&'):
        super().__init__(jointype.join(vs.name for vs in sheets))
        self.source = sheets
        self.jointype = jointype

    def reload(self):
        super().reload()
        sheets = self.source
        # first element in joined row is the tuple of keys
        self.columns = []
        for colnum in range(sheets[0].nKeys):
            c = sheets[0].columns[colnum]
            self.columns.append(Column(c.name, c.type, lambda_subrow_wrap(lambda_col(colnum), 0)))
        self.nKeys = sheets[0].nKeys

        rowsBySheetKey = {}
        rowsByKey = {}

        for vs in sheets:
            rowsBySheetKey[vs] = {}
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                rowsBySheetKey[vs][key] = r

        for sheetnum, vs in enumerate(sheets):
            # subsequent elements are the rows from each source, in order of the source sheets
            self.columns.extend(Column(c.name, c.type, lambda_subrow_wrap(c.func, sheetnum+1), c.width) for c in vs.columns[vs.nKeys:])
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                if key not in rowsByKey:
                    rowsByKey[key] = [key] + [rowsBySheetKey[vs2].get(key) for vs2 in sheets]  # combinedRow

        if self.jointype == '&':  # inner join  (only rows with matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if all(combinedRow))
        elif self.jointype == '+':  # outer join (all rows from first sheet)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if combinedRow[1])
        elif self.jointype == '*':  # full join (keep all rows from all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items())
        elif self.jointype == '~':  # diff join (only rows without matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if not all(combinedRow))

class SheetAppend(Sheet):
    'rows are appended; union of all columns (same names are assumed the same if collapse_columns)'
    def __init__(self, sheets, collapse_columns=True):
        super().__init__('+'.join(str(vs) for vs in sheets), sheets)
        self.collapse_columns = collapse_columns

    def reload(self):
        self.rows = []
        self.columns = []
        for i, vs in enumerate(self.source):
            self.rows.extend(zip([i]*len(vs.rows), vs.rows))
            self.columns.extend(Column(c.name, c.type, lambda_if_firstcol(c.func, i), c.width) for c in vs.columns)
#            if self.collapse_columns:
#                    Multiplexer([col[[lambda_col(i) for i,c in enumerate(self.source[0].columns)]

class SheetFreqTable(Sheet):
    def __init__(self, sheet, col):
        fqcolname = '%s_%s_freq' % (sheet.name, col.name)
        super().__init__(fqcolname, sheet)

        self.origCol = col
        self.values = collections.defaultdict(list)
        for r in sheet.rows:
            self.values[str(col.getValue(r))].append(r)

    def reload(self):
        super().reload()
        self.rows = sorted(self.values.items(), key=lambda r: len(r[1]), reverse=True)  # sort by num reverse
        self.largest = len(self.rows[0][1])+1

        self.columns = [
            ColumnItem(self.origCol.name, 0, type=self.origCol.type),
            Column('num', int, lambda r: len(r[1])),
            Column('percent', float, lambda r: len(r[1])*100/self.source.nRows),
            Column('histogram', str, lambda r,s=self: options.ch_Histogram*int(len(r[1])*80/s.largest), width=80)
        ]
        self.nKeys = 1

        self.command(Key(' '), 'source.toggle(cursorRow[1])', 'toggle these entries')
        self.command(Key('s'), 'source.select(cursorRow[1])', 'select these entries')
        self.command(Key('u'), 'source.unselect(cursorRow[1])', 'unselect these entries')
        self.command(Key.ENTER, 'vd.push(source.copy()).rows = values[str(columns[0].getValue(cursorRow))]', 'push new sheet with only this value')


### input formats and helpers

sourceCache = {}

def getTextContents(p):
    if not p in sourceCache:
        sourceCache[p] = p.read_text(encoding=options.encoding, errors=options.encoding_errors)
    return sourceCache[p]

def open_source(p, filetype=None):
    if isinstance(p, Path):
        if filetype is None:
            filetype = p.suffix

        if p.is_dir():
            vs = SheetDirectory(p)
        else:
            openfunc = 'open_' + filetype
            if openfunc not in globals():
                openfunc = 'open_txt'
                status('no %s function' % openfunc)
            vs = globals()[openfunc](p)
    elif isinstance(p, str):
        if '://' in p:
            vs = openUrl(p)
        else:
            return open_source(Path(p), filetype)
    else:  # some other object
        status('unknown object type %s' % type(p))
        vs = None

    if vs:
        status('opened %s' % p.name)
    return vs

class SheetText(Sheet):
    def __init__(self, name, content, src=None):
        super().__init__(name, src)
        self.content = content

    def reload(self):
        if isinstance(self.content, list):
            self.rows = self.content
        elif isinstance(self.content, str):
            self.rows = self.content.split('\n')
        else:
            error('unknown text type ' + str(type(self.content)))
        self.columns = [Column(self.name, str)]

class SheetDirectory(Sheet):
    def __init__(self, p):
        super().__init__(p.name, p)

    def reload(self):
        self.command(Key.ENTER, 'vd.push(open_source(cursorRow[0]))', 'open file')  # path, filename
        self.reload_commands()
        self.rows = [(p, p.stat()) for p in self.source.iterdir()]  #  if not p.name.startswith('.')]
        self.columns = [Column('filename', str, lambda r: r[0].name),
                        Column('type', str, lambda r: r[0].is_dir() and '/' or r[0].suffix),
                        Column('size', int, lambda r: r[1].st_size),
                        Column('mtime', date, lambda r: r[1].st_mtime)]

def open_txt(p):
    contents = getTextContents(p)
    if '\t' in contents[:32]:
        return open_tsv(p)  # TSV often have .txt extension
    return SheetText(p.name, contents, p)

def open_csv(p):
    contents = getTextContents(p)
    if options.csv_dialect == 'sniff':
        headers = contents[:1024]
        dialect = csv.Sniffer().sniff(headers)
        status('sniffed csv_dialect as %s' % dialect)
    else:
        dialect = options.csv_dialect

    return load_csv(Sheet(p.name, p), contents,
                        header=options.csv_header,
                        dialent=dialect,
                        quotechar=options.csv_quotechar,
                        delimiter=ptions.csv_delimiter)

def load_csv(vs, contents, header=False, **kwargs):
    rdr = csv.reader(io.StringIO(contents, newline=''), **kwargs)
    vs.rows = [r for r in rdr]
    if header:
        vs.columns = ArrayNamedColumns(vs.rows[0])
        vs.rows = vs.rows[1:]
    else:
        vs.columns = ArrayColumns(len(vs.rows[0]))

    return vs

def open_tsv(p, **kwargs):
    return load_tsv(Sheet(p.name, p), getTextContents(p), header=options.csv_header)

def load_tsv(vs, contents, header=False):
    'populates vs with the parsed tsv text in contents.'
    lines = contents.splitlines()
    if lines:
        toprow = lines[0].split('\t')

        if header:
            vs.columns = ArrayNamedColumns(toprow)
            lines = lines[1:]
        else:
            vs.columns = ArrayColumns(len(toprow))

        vs.rows = [L.split('\t') for L in lines]  # [rownum] -> [ field, ... ]
    return vs

def open_json(p):
    import json
    return load_pyobj(p.name, json.loads(getTextContents(p)))

#### .xlsx
class open_xlsx(Sheet):
    def __init__(self, path):
        super().__init__(path.name, path)
        import openpyxl
        self.workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        self.rows = list(self.workbook.sheetnames)
        self.columns = [Column('name', str)]
        self.command(Key.ENTER, 'vd.push(sheet.getSheet(cursorRow))', 'push this sheet')

    def getSheet(self, sheetname):
        'create actual Sheet from xlsx sheet'
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        vs = Sheet('%s%s%s' % (self.source, options.SubsheetSep, sheetname), worksheet)
        vs.columns = ArrayColumns(worksheet.max_column)
        vs.rows = [ [cell.value for cell in row] for row in worksheet.iter_rows()]
        return vs

#### .hdf5
class SheetH5Obj(Sheet):
    def __init__(self, name, hobj, src):
        super().__init__(name, src)
        self.hobj = hobj

    def reload(self):
        super().reload()
        import h5py
        if isinstance(self.hobj, h5py.Group):
            self.rows = [ self.hobj[objname] for objname in self.hobj.keys() ]
            self.columns = [
                Column(self.hobj.name, str, lambda r: r.name.split('/')[-1]),
                Column('type', str, lambda r: type(r).__name__),
                Column('nItems', int, lambda r: len(r)),
            ]
            self.command(Key.ENTER, 'vd.push(SheetH5Obj(name+options.SubsheetSep+cursorRow.name, cursorRow, source))', 'open this group or dataset')
            self.command(Key('A'), 'vd.push(SheetDict(cursorRow.name + "_attrs", cursorRow.attrs))', 'open metadata sheet for this object')
        elif isinstance(self.hobj, h5py.Dataset):
            if len(self.hobj.shape) == 1:
                self.rows = self.hobj[:]  # copy
                self.columns = [Column(colname, str, lambda_colname(colname)) for colname in self.hobj.dtype.names]
            elif len(self.hobj.shape) == 2:  # matrix
                self.rows = self.hobj[:]  # copy
                self.columns = ArrayColumns(self.hobj.shape[1])
            else:
                status('too many dimensions in shape %s' % str(self.hobj.shape))
        else:
            status('unknown h5 object type %s' % type(self.hobj))

class open_hdf5(SheetH5Obj):
    def __init__(self, p):
        import h5py
        super().__init__(p.name, h5py.File(str(p), 'r'), p)
open_h5 = open_hdf5


class open_zip(Sheet):
    def __init__(self, p):
        import zipfile
        super().__init__(p.name, p)
        self.zfp = zipfile.ZipFile(p.fqpn, 'r')
        self.rows = self.zfp.infolist()
        self.columns = AttrColumns("filename file_size date_time compress_size".split())
        self.command(Key.ENTER, 'vd.push(open(cursorRow))', 'open this file')

    def open(self, zi):
        cachefn = zi.filename
        if not os.path.exists(cachefn):
            self.zfp.extract(zi)
        return open_source(cachefn)

#### databases
class SheetBlaze(Sheet):
    def __init__(self, name, data, src):
        super().__init__(name, src)
        self.columns = ArrayNamedColumns(data.fields)
        self.rows = list(data)

def openUrl(url):
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if m:
        return open_gspreadsheet(Path(m.group(1)))

    import blaze
    import datashape; datashape.coretypes._canonical_string_encodings.update({"utf8_unicode_ci": "U8"})
    fp = blaze.data(url)
    vs = SheetList(url, [getattr(fp, tblname) for tblname in fp.fields], url)
    vs.command(Key.ENTER, 'vd.push(SheetBlaze(cursorRow.name, cursorRow, sheet))', 'open this table')
    return vs

#### Google Sheets; requires credentials to be setup already

@functools.lru_cache()
def google_sheets():
    import httplib2
    import os

    from apiclient import discovery
    from oauth2client.file import Storage

    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    credential_path = os.path.join(credential_dir, 'sheets.googleapis.com-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials:
        status('Credentials required')
    elif credentials.invalid:
        status('No or Invalid credentials')
    else:
        http = credentials.authorize(httplib2.Http())
        discoveryUrl = ('https://sheets.googleapis.com/$discovery/rest?version=v4')
        service = discovery.build('sheets', 'v4', http=http, discoveryServiceUrl=discoveryUrl)

        return service.spreadsheets()

def open_gspreadsheet(p):
    sheets = google_sheets()
    sheet_md = sheets.get(spreadsheetId=p.name).execute()
    vs = Sheet(sheet_md['properties']['title'], p)
    vs.columns = [Column('title', lambda_eval('properties.title')),
                  Column('rowCount', lambda_eval('properties.gridProperties.rowCount')),
                  Column('columnCount', lambda_eval('properties.gridProperties.columnCount'))]
    vs.rows = sheet_md['sheets']
    vs.command(Key.ENTER, 'cursorRow')
    return vs

def open_gsheet(p):
    sheets = google_sheets()
    sheet = sheets.values().get(spreadsheetId=p.name).execute()
    push_pyobj(p.name, sheet, p)
#    vs = Sheet(sheet_md['properties']['title'], p)
#    vs.columns = [Column('title', lambda_eval('properties.title')),
#                  Column('rowCount', lambda_eval('properties.gridProperties.rowCount')),
#                  Column('columnCount', lambda_eval('properties.gridProperties.columnCount'))]
#    vs.rows = sheet_md['sheets']
#    vs.command(Key.ENTER, 'cursorRow')
#    return vs

#### external addons
def open_py(p):
    contents = getTextContents(p)
    exec(contents, globals())
    status('executed %s' % p)

### Sheet savers

def saveSheet(sheet, fn):
    if options.readonly:
        status('readonly mode')
        return
    basename, ext = os.path.splitext(fn)
    funcname = 'save_' + ext[1:]
    globals().get(funcname, save_tsv)(sheet, fn)
    status('saved to ' + fn)

def save_tsv(sheet, fn):
    with open(fn, 'w', encoding=options.encoding, errors=options.encoding_errors) as fp:
        colhdr = '\t'.join(col.name for col in sheet.visibleCols) + '\n'
        if colhdr.strip():  # is anything but whitespace
            fp.write(colhdr)
        for r in sheet.rows:
            fp.write('\t'.join(col.getDisplayValue(r) for col in sheet.visibleCols) + '\n')

def save_csv(sheet, fn):
    with open(fn, 'w', newline='', encoding=options.encoding, errors=options.encoding_errors) as fp:
        cw = csv.writer(fp, dialect=options.csv_dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        colnames = [col.name for col in sheet.visibleCols]
        if ''.join(colnames):
            cw.writerow(colnames)
        for r in sheet.rows:
            cw.writerow([col.getDisplayValue(r) for col in sheet.visibleCols])

### curses, options, init

def inputLine(prompt, value='', completions=None):
    'add a prompt to the bottom of the screen and get a line of input from the user'
    scr = vd().scr
    windowHeight, windowWidth = scr.getmaxyx()
    scr.addstr(windowHeight-1, 0, prompt)
    return vd().editText(windowHeight-1, len(prompt), windowWidth-len(prompt)-8, value=value, attr=colors[options.c_EditCell], unprintablechar=options.ch_Unprintable, completions=completions)


def run(sheetlist=None):
    os.putenv('ESCDELAY', '25')  # reduce ESC timeout to 25ms. http://en.chys.info/2009/09/esdelay-ncurses/
    ret = wrapper(curses_main, sheetlist)
    if ret:
        print(ret)

def curses_main(_scr, sheetlist=None):
    try:
        setup_sheet_commands()
        if sheetlist:
            for vs in sheetlist:
                vd().push(vs)
        return vd().run(_scr)
    except Exception as e:
        if options.debug:
            raise
        return '%s: %s' % (type(e).__name__, str(e))



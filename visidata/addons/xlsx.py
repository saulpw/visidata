
from visidata import *

class open_xlsx(Sheet):
    """Load XLSX file (in Excel Open XML format)."""
    def __init__(self, path):
        super().__init__(path.name, path)
        self.workbook = None
        self.command(ENTER, 'vd.push(sheet.getSheet(cursorRow))', 'push this sheet')

    @async
    def reload(self):
        import openpyxl
        self.columns = [Column('name')]
        self.workbook = openpyxl.load_workbook(self.source.resolve(), data_only=True, read_only=True)
        self.rows = list(self.workbook.sheetnames)

    def getSheet(self, sheetname):
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        return xlsxSheet(joinSheetnames(self.name, sheetname), worksheet)

class xlsxSheet(Sheet):
    @async
    def reload(self):
        worksheet = self.source
        self.columns = ArrayColumns(worksheet.max_column)
        self.progressTotal = worksheet.max_row
        for row in worksheet.iter_rows():
            self.progressMade += 1
            self.rows.append([cell.value for cell in row])

class open_xls(Sheet):
    """Load XLS file (in Excel format)."""
    def __init__(self, path):
        super().__init__(path.name, path)
        self.workbook = None
        self.command(ENTER, 'vd.push(sheet.getSheet(cursorRow))', 'push this sheet')

    @async
    def reload(self):
        import xlrd
        self.columns = [Column('name')]
        self.workbook = xlrd.open_workbook(self.source.resolve())
        self.rows = list(self.workbook.sheet_names())

    def getSheet(self, sheetname):
        worksheet = self.workbook.sheet_by_name(sheetname)
        return xlsSheet(joinSheetnames(self.name, sheetname), worksheet)


class xlsSheet(Sheet):
    @async
    def reload(self):
        worksheet = self.source
        self.columns = ArrayColumns(worksheet.ncols)
        self.progressTotal = worksheet.nrows
        for rownum in range(worksheet.nrows):
            self.rows.append([worksheet.cell(rownum, colnum).value for colnum in range(worksheet.ncols)])
            self.progressMade += 1

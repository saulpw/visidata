import csv
import io
from . import status, option, options
from . import FileBrowser, TextViewer, join_sheetnames
from .columns import Column, ArrayColumns, ArrayNamedColumns, ColumnItem
from .sheets import Sheet
from .Path import Path

### input formats and helpers

sourceCache = {}

def getTextContents(p):
    if not p in sourceCache:
        sourceCache[p] = p.read_text(encoding=options.encoding, errors=options.encoding_errors)
    return sourceCache[p]

def open_source(p, filetype=None):
    if isinstance(p, Path):
        if filetype is None:
            filetype = p.suffix

        if p.is_dir():
            vs = FileBrowser(p)
        else:
            openfunc = 'open_' + filetype
            if openfunc not in globals():
                openfunc = 'open_txt'
                status('no %s function' % openfunc)
            vs = globals()[openfunc](p)
    elif isinstance(p, str):
        if '://' in p:
            vs = openUrl(p)
        else:
            return open_source(Path(p), filetype)
    else:  # some other object
        status('unknown object type %s' % type(p))
        vs = None

    if vs:
        status('opened %s' % p.name)
    return vs

def open_txt(p):
    contents = getTextContents(p)
    if '\t' in contents[:32]:
        return open_tsv(p)  # TSV often have .txt extension
    return TextViewer(p.name, contents, p)

def open_json(p):
    import json
    return load_pyobj(p.name, json.loads(getTextContents(p)))

#### external addons
def open_py(p):
    contents = getTextContents(p)
    exec(contents, globals())
    status('executed %s' % p)

## csv/tsv

option('csv_dialect', 'excel', 'dialect passed to csv.reader')
option('csv_delimiter', ',', 'delimiter passed to csv.reader')
option('csv_quotechar', '"', 'quotechar passed to csv.reader')
option('csv_headerlines', '0', 'parse first row of CSV as column names')

option('encoding', 'utf-8', 'as passed to codecs.open')
option('encoding_errors', 'surrogateescape', 'as passed to codecs.open')

def open_csv(p):
    vs = Sheet(p.name, p)
    vs.contents = getTextContents(p)
    vs.loader = lambda vs=vs: load_csv(vs)
    return vs

def load_csv(vs):
    header_lines = int(options.csv_headerlines or 0)
    if options.csv_dialect == 'sniff':
        headers = self.contents[:1024]
        dialect = csv.Sniffer().sniff(headers)
        status('sniffed csv_dialect as %s' % dialect)
    else:
        dialect = options.csv_dialect

    rdr = csv.reader(io.StringIO(vs.contents, newline=''), 
                        dialect=dialect,
                        quotechar=options.csv_quotechar,
                        delimiter=options.csv_delimiter)
    rows = [r for r in rdr]
    if header_lines:
        # columns ideally reflect the max number of fields over all rows
        vs.columns = ArrayNamedColumns('\\n'.join(x) for x in zip(*rows[:header_lines]))
    else:
        vs.columns = ArrayColumns(len(rows[0]))
    vs.rows = rows[header_lines:]
    return vs


def open_tsv(p):
    vs = Sheet(p.name, p)
    vs.contents = getTextContents(p)
    vs.loader = lambda vs=vs: load_tsv(vs)
    return vs

# separate function, so reloads can be triggered
def load_tsv(vs):
    'populates vs with the parsed tsv text in contents.'
    header_lines = int(options.csv_headerlines or 0)
    lines = vs.contents.splitlines()
    if lines:
        rows = [L.split('\t') for L in lines]
        if header_lines:
            # columns ideally reflect the max number of fields over all rows
            vs.columns = ArrayNamedColumns('\\n'.join(x) for x in zip(*rows[:header_lines]))
        else:
            vs.columns = ArrayColumns(len(rows[0]))

        vs.rows = rows[header_lines:]
        status('lines (lines=%d, )' % len(vs.columns))
    else:
        status('no lines (len %d)' % len(vs.columns))
    return vs

def save_tsv(sheet, fn):
    with open(fn, 'w', encoding=options.encoding, errors=options.encoding_errors) as fp:
        colhdr = '\t'.join(col.name for col in sheet.visibleCols) + '\n'
        if colhdr.strip():  # is anything but whitespace
            fp.write(colhdr)
        for r in sheet.rows:
            fp.write('\t'.join(col.getDisplayValue(r) for col in sheet.visibleCols) + '\n')

def save_csv(sheet, fn):
    with open(fn, 'w', newline='', encoding=options.encoding, errors=options.encoding_errors) as fp:
        cw = csv.writer(fp, dialect=options.csv_dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        colnames = [col.name for col in sheet.visibleCols]
        if ''.join(colnames):
            cw.writerow(colnames)
        for r in sheet.rows:
            cw.writerow([col.getDisplayValue(r) for col in sheet.visibleCols])

## xlsx

class open_xlsx(Sheet):
    def __init__(self, path):
        super().__init__(path.name, path)
        import openpyxl
        self.workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    def reload(self):
        self.rows = list(self.workbook.sheetnames)
        self.columns = [Column('name')]
        self.command('^J', 'vd.push(sheet.getSheet(cursorRow))', 'push this sheet')

    def getSheet(self, sheetname):
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        return xlsxSheet(join_sheetnames(self.source, sheetname), worksheet)

class xlsxSheet(Sheet):
    def reload(self):
        worksheet = self.source
        self.columns = ArrayColumns(worksheet.max_column)
        self.rows = [[cell.value for cell in row] for row in worksheet.iter_rows()]

## hdf5

class SheetH5Obj(Sheet):
    def __init__(self, name, hobj, src):
        super().__init__(name, src)
        self.hobj = hobj

    def reload(self):
        super().reload()
        import h5py
        if isinstance(self.hobj, h5py.Group):
            self.rows = [ self.hobj[objname] for objname in self.hobj.keys() ]
            self.columns = [
                Column(self.hobj.name, str, lambda r: r.name.split('/')[-1]),
                Column('type', str, lambda r: type(r).__name__),
                Column('nItems', int, lambda r: len(r)),
            ]
            self.command('^J', 'vd.push(SheetH5Obj(name+options.SubsheetSep+cursorRow.name, cursorRow, source))', 'open this group or dataset')
            self.command('A', 'vd.push(SheetDict(cursorRow.name + "_attrs", cursorRow.attrs))', 'open metadata sheet for this object')
        elif isinstance(self.hobj, h5py.Dataset):
            if len(self.hobj.shape) == 1:
                self.rows = self.hobj[:]  # copy
                self.columns = [ColumnItem(colname, colname) for colname in self.hobj.dtype.names]
            elif len(self.hobj.shape) == 2:  # matrix
                self.rows = self.hobj[:]  # copy
                self.columns = ArrayColumns(self.hobj.shape[1])
            else:
                status('too many dimensions in shape %s' % str(self.hobj.shape))
        else:
            status('unknown h5 object type %s' % type(self.hobj))

class open_hdf5(SheetH5Obj):
    def __init__(self, p):
        import h5py
        super().__init__(p.name, h5py.File(str(p), 'r'), p)

open_h5 = open_hdf5
